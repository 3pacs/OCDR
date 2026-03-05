"""
Flexible Excel Import Engine.

Handles ANY Excel file — auto-detects headers, fuzzy-matches columns to the
billing_records schema, streams large files in chunks, and stores unmapped
columns in the extra_data JSONB field so no data is lost.

Designed for 100MB+ files with messy, inconsistent column names.
"""

import io
import logging
from datetime import date, datetime, timedelta

from openpyxl import load_workbook
from rapidfuzz import fuzz, process
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.app.models.billing import BillingRecord
from backend.app.models.import_file import ImportFile

logger = logging.getLogger(__name__)

BATCH_SIZE = 500
EXCEL_EPOCH = date(1899, 12, 30)

# ─── Column mapping aliases ──────────────────────────────────────────────────
# Each key is the BillingRecord field name, values are known header aliases
# sorted by priority. Fuzzy matching will try these + the field name itself.
COLUMN_ALIASES = {
    "patient_name": [
        "patient", "patient name", "patient_name", "name", "pt name",
        "pt", "client name", "client", "patientname",
    ],
    "referring_doctor": [
        "doctor", "referring doctor", "ref doctor", "ref doc", "physician",
        "referring physician", "ordering doctor", "ordering physician",
        "ref", "dr", "referring_doctor",
    ],
    "scan_type": [
        "scan", "scan type", "scan_type", "exam", "exam type", "procedure",
        "procedure type", "study", "study type",
    ],
    "gado_used": [
        "gado", "gadolinium", "contrast", "gad", "gado_used", "gado used",
        "contrast used", "with contrast",
    ],
    "insurance_carrier": [
        "insurance", "carrier", "insurance carrier", "insurance_carrier",
        "payer", "payor", "ins", "ins carrier", "insurer", "plan",
        "insurance company", "insurance name",
    ],
    "modality": [
        "modality", "type", "mod", "imaging type", "scan modality",
        "equipment", "machine type",
    ],
    "service_date": [
        "date", "service date", "service_date", "dos", "date of service",
        "exam date", "study date", "visit date", "svc date", "appt date",
    ],
    "primary_payment": [
        "primary", "primary payment", "primary_payment", "primary pay",
        "pri payment", "pri pay", "1st payment",
    ],
    "secondary_payment": [
        "secondary", "secondary payment", "secondary_payment", "secondary pay",
        "sec payment", "sec pay", "2nd payment",
    ],
    "total_payment": [
        "total", "total payment", "total_payment", "total paid", "amount paid",
        "payment", "paid", "amount", "pay amount", "total pay",
    ],
    "extra_charges": [
        "extra", "extra charges", "extra_charges", "additional charges",
        "add charges", "misc charges", "other charges", "surcharge",
    ],
    "reading_physician": [
        "reading physician", "reading_physician", "read by", "readby",
        "radiologist", "interpreting", "interpreter", "reading dr",
        "read", "interp",
    ],
    "patient_id": [
        "patient id", "patient_id", "id", "mrn", "medical record",
        "record number", "chart number", "acct", "account",
        "patient number", "patid", "pt id",
    ],
    "birth_date": [
        "birth date", "birth_date", "dob", "date of birth", "birthdate",
        "birthday",
    ],
    "patient_name_display": [
        "patient name display", "display name", "full name",
        "patient_name_display", "formatted name",
    ],
    "schedule_date": [
        "schedule date", "scheduled date", "s date", "sched date",
        "schedule_date", "appointment date",
    ],
    "modality_code": [
        "modality code", "modality_code", "modalities", "mod code",
    ],
    "description": [
        "description", "desc", "exam description", "procedure description",
        "study description", "notes", "comments", "detail",
    ],
    "service_month": [
        "month", "service month", "service_month", "svc month",
    ],
    "service_year": [
        "year", "service year", "service_year", "svc year",
    ],
    "is_new_patient": [
        "new", "new patient", "is_new_patient", "is new", "new pt",
    ],
}

# Fields that accept money values
MONEY_FIELDS = {"primary_payment", "secondary_payment", "total_payment", "extra_charges"}
# Fields that accept date values
DATE_FIELDS = {"service_date", "birth_date", "schedule_date"}
# Fields that accept boolean values
BOOL_FIELDS = {"gado_used", "is_new_patient"}
# Fields that accept integer values
INT_FIELDS = {"patient_id"}


def _excel_serial_to_date(serial) -> date | None:
    if serial is None:
        return None
    if isinstance(serial, datetime):
        return serial.date()
    if isinstance(serial, date):
        return serial
    try:
        serial = int(float(serial))
        if serial < 1:
            return None
        return EXCEL_EPOCH + timedelta(days=serial)
    except (ValueError, TypeError, OverflowError):
        return None


def _parse_date_value(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Try Excel serial
    d = _excel_serial_to_date(val)
    if d:
        return d
    # Try common string formats
    s = str(val).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_money(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    s = str(val).strip().replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return 0.0


def _parse_bool(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("YES", "TRUE", "1", "Y", "X")


def _clean_text(val, max_len: int = 200) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    return s[:max_len]


SELFPAY_VARIANTS = {"SELFPAY", "SELF-PAY", "SELF PAY", "CASH", "SELF"}


def _normalize_carrier(carrier: str | None) -> str | None:
    if not carrier:
        return carrier
    upper = carrier.upper()
    if upper in SELFPAY_VARIANTS:
        return "SELF PAY"
    return upper


def _derive_psma(description: str | None, modality: str | None) -> bool:
    if not description:
        return False
    desc_upper = description.upper()
    if "PSMA" in desc_upper:
        return True
    if modality and modality.upper() == "PET" and ("GA-68" in desc_upper or "GALLIUM" in desc_upper):
        return True
    return False


# ─── Header Detection ────────────────────────────────────────────────────────

def _detect_header_row(ws, max_scan: int = 20) -> tuple[int, list[str]]:
    """
    Scan the first N rows to find the header row.
    Heuristic: the row with the most non-empty text cells that look like headers.
    Returns (row_number_1based, list_of_header_strings).
    """
    best_row = 1
    best_headers = []
    best_score = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        if row is None:
            continue
        headers = []
        text_count = 0
        for cell in row:
            if cell is not None:
                s = str(cell).strip()
                headers.append(s)
                # Score: prefer text cells, penalize purely numeric
                if s and not s.replace(".", "").replace("-", "").isdigit():
                    text_count += 1
                else:
                    headers[-1] = s  # keep it but don't count
            else:
                headers.append("")

        if text_count > best_score:
            best_score = text_count
            best_row = row_idx
            best_headers = headers

    return best_row, best_headers


def _fuzzy_map_columns(headers: list[str], threshold: int = 60) -> tuple[dict[int, str], list[str]]:
    """
    Map column indices to BillingRecord field names using fuzzy matching.
    Returns (mapping: {col_idx: field_name}, unmapped: [header_names]).
    """
    # Build a lookup: alias -> field_name
    alias_to_field = {}
    all_aliases = []
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_to_field[alias.lower()] = field
            all_aliases.append(alias.lower())

    mapping = {}
    used_fields = set()
    unmapped = []

    for col_idx, raw_header in enumerate(headers):
        if not raw_header:
            continue
        header_lower = raw_header.lower().strip()

        # Exact match first
        if header_lower in alias_to_field:
            field = alias_to_field[header_lower]
            if field not in used_fields:
                mapping[col_idx] = field
                used_fields.add(field)
                continue

        # Fuzzy match
        match = process.extractOne(header_lower, all_aliases, scorer=fuzz.WRatio, score_cutoff=threshold)
        if match:
            field = alias_to_field[match[0]]
            if field not in used_fields:
                mapping[col_idx] = field
                used_fields.add(field)
                continue

        unmapped.append(raw_header)

    return mapping, unmapped


# ─── Sheet inspection ────────────────────────────────────────────────────────

def inspect_excel_file(file_content: bytes) -> dict:
    """
    Quick scan of an Excel file: return sheet names, detected headers per sheet,
    and proposed column mappings. Used for the preview UI before full import.
    """
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, headers = _detect_header_row(ws)
        mapping, unmapped = _fuzzy_map_columns(headers)

        # Convert mapping to readable format
        readable_mapping = {}
        for col_idx, field in mapping.items():
            readable_mapping[headers[col_idx]] = field

        row_count = 0
        for _ in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_count += 1

        sheets[sheet_name] = {
            "header_row": header_row,
            "headers": [h for h in headers if h],
            "total_columns": len([h for h in headers if h]),
            "mapped_columns": readable_mapping,
            "unmapped_columns": unmapped,
            "estimated_rows": row_count,
        }
    wb.close()
    return {"sheets": sheets, "sheet_names": wb.sheetnames}


# ─── String truncation safety ────────────────────────────────────────────────

# Max lengths matching the model's VARCHAR sizes
FIELD_MAX_LENGTHS = {
    "patient_name": 200,
    "referring_doctor": 200,
    "scan_type": 200,
    "insurance_carrier": 200,
    "modality": 100,
    "reading_physician": 200,
    "patient_name_display": 200,
    "modality_code": 100,
    "service_month": 20,
    "service_year": 10,
    "denial_status": 50,
    "denial_reason_code": 50,
    "era_claim_id": 50,
    "import_source": 50,
}


def _truncate_fields(record_data: dict) -> None:
    """Truncate string fields to their column max lengths to prevent DB errors."""
    for field, max_len in FIELD_MAX_LENGTHS.items():
        if field in record_data and isinstance(record_data[field], str):
            record_data[field] = record_data[field][:max_len]


# ─── Main import ─────────────────────────────────────────────────────────────

async def import_excel_flexible(
    file_content: bytes,
    filename: str,
    session: AsyncSession,
    sheet_name: str | None = None,
    column_overrides: dict[str, str] | None = None,
) -> dict:
    """
    Import any Excel file into billing_records with smart column detection.

    - Auto-detects header row
    - Fuzzy-matches columns to schema
    - Stores unmapped columns in extra_data JSONB
    - Deduplicates on patient+date+scan+modality
    - Handles 100MB+ files via read_only mode

    Args:
        file_content: Raw bytes of the Excel file
        filename: Original filename for tracking
        session: DB session
        sheet_name: Specific sheet to import (None = first sheet)
        column_overrides: Manual overrides {header_name: field_name}

    Returns: dict with import stats
    """
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)

    # Pick sheet
    if sheet_name and sheet_name in wb.sheetnames:
        target_sheet = sheet_name
    elif sheet_name:
        # Fuzzy match sheet name
        match = process.extractOne(sheet_name, wb.sheetnames, scorer=fuzz.WRatio, score_cutoff=50)
        if match:
            target_sheet = match[0]
        else:
            available = ", ".join(wb.sheetnames)
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {available}")
    else:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]

    # Detect headers
    header_row, headers = _detect_header_row(ws)
    mapping, unmapped_cols = _fuzzy_map_columns(headers)

    # Apply manual overrides
    if column_overrides:
        for header_name, field_name in column_overrides.items():
            for col_idx, h in enumerate(headers):
                if h.lower().strip() == header_name.lower().strip():
                    mapping[col_idx] = field_name
                    if header_name in unmapped_cols:
                        unmapped_cols.remove(header_name)

    # Track unmapped column indices for extra_data
    unmapped_indices = {}
    for col_idx, h in enumerate(headers):
        if col_idx not in mapping and h:
            unmapped_indices[col_idx] = h

    # Create import file record
    import_file = ImportFile(
        filename=filename,
        file_size_bytes=len(file_content),
        sheet_name=target_sheet,
        import_type="EXCEL_FLEXIBLE",
        status="PROCESSING",
        column_mapping={headers[k]: v for k, v in mapping.items()},
        unmapped_columns=unmapped_cols,
    )
    session.add(import_file)
    await session.flush()

    # Load existing dedup keys
    existing_result = await session.execute(
        select(
            BillingRecord.patient_name,
            BillingRecord.service_date,
            BillingRecord.scan_type,
            BillingRecord.modality,
        )
    )
    existing_keys = {(r[0], r[1], r[2], r[3]) for r in existing_result.fetchall()}

    imported = 0
    skipped = 0
    errors = 0
    batch: list[BillingRecord] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        try:
            if row is None or all(c is None for c in row):
                skipped += 1
                continue

            # Build record from mapped columns
            record_data = {}
            extra_data = {}

            for col_idx, field in mapping.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]

                if field in DATE_FIELDS:
                    record_data[field] = _parse_date_value(val)
                elif field in MONEY_FIELDS:
                    record_data[field] = _parse_money(val)
                elif field in BOOL_FIELDS:
                    record_data[field] = _parse_bool(val)
                elif field in INT_FIELDS:
                    if val is not None:
                        try:
                            record_data[field] = int(float(val))
                        except (ValueError, TypeError):
                            record_data[field] = None
                    else:
                        record_data[field] = None
                else:
                    record_data[field] = _clean_text(val)

            # Collect unmapped columns into extra_data
            for col_idx, header_name in unmapped_indices.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    val = row[col_idx]
                    # Convert dates/datetimes to strings for JSON
                    if isinstance(val, (datetime, date)):
                        val = val.isoformat()
                    elif not isinstance(val, (str, int, float, bool)):
                        val = str(val)
                    extra_data[header_name] = val

            # Validate minimum fields — must have at least patient name
            patient_name = record_data.get("patient_name")
            if not patient_name:
                skipped += 1
                continue

            # Normalize carrier
            if "insurance_carrier" in record_data:
                record_data["insurance_carrier"] = _normalize_carrier(record_data["insurance_carrier"])

            # Derive PSMA
            description = record_data.get("description")
            modality = record_data.get("modality")
            record_data["is_psma"] = _derive_psma(description, modality)

            # Dedup check (only if we have all 4 dedup fields)
            service_date = record_data.get("service_date")
            scan_type = record_data.get("scan_type")
            if patient_name and service_date and scan_type and modality:
                dedup_key = (patient_name.upper(), service_date, scan_type.upper() if scan_type else None, modality.upper() if modality else None)
                if dedup_key in existing_keys:
                    skipped += 1
                    continue
                existing_keys.add(dedup_key)

            # Set import metadata
            record_data["import_source"] = f"FLEX:{filename[:30]}"
            record_data["import_file_id"] = import_file.id
            if extra_data:
                record_data["extra_data"] = extra_data

            # Fill in required fields with defaults if missing
            if "referring_doctor" not in record_data or not record_data["referring_doctor"]:
                record_data["referring_doctor"] = "UNKNOWN"
            if "scan_type" not in record_data or not record_data["scan_type"]:
                record_data["scan_type"] = "UNKNOWN"
            if "insurance_carrier" not in record_data or not record_data["insurance_carrier"]:
                record_data["insurance_carrier"] = "UNKNOWN"
            if "modality" not in record_data or not record_data["modality"]:
                record_data["modality"] = "UNKNOWN"
            if "service_date" not in record_data or not record_data["service_date"]:
                record_data["service_date"] = date.today()

            # Truncate all string fields to their column max lengths
            _truncate_fields(record_data)

            # Validate before inserting
            from backend.app.analytics.data_validation import validate_billing_record
            validation_errors = [v for v in validate_billing_record(record_data) if v.severity == "ERROR"]
            if validation_errors:
                logger.warning(f"Row {row_idx} validation: {validation_errors[0].message}")
                errors += 1
                continue

            batch.append(BillingRecord(**record_data))
            imported += 1

            if len(batch) >= BATCH_SIZE:
                try:
                    session.add_all(batch)
                    await session.flush()
                except Exception as e:
                    logger.warning(f"Batch flush error at row {row_idx}, rolling back batch: {e}")
                    await session.rollback()
                    errors += len(batch)
                    imported -= len(batch)
                batch = []

        except Exception as e:
            logger.warning(f"Row {row_idx} error: {e}")
            errors += 1

    # Flush remaining
    if batch:
        try:
            session.add_all(batch)
            await session.flush()
        except Exception as e:
            logger.warning(f"Final batch flush error, rolling back: {e}")
            await session.rollback()
            errors += len(batch)
            imported -= len(batch)

    # Update import file record — re-merge in case of prior rollback
    try:
        import_file = await session.merge(import_file)
        import_file.status = "COMPLETED" if errors == 0 else "COMPLETED_WITH_ERRORS"
        import_file.rows_imported = imported
        import_file.rows_skipped = skipped
        import_file.rows_errored = errors
        import_file.completed_at = datetime.utcnow()
        await session.commit()
    except Exception as e:
        logger.warning(f"Could not update import_file record: {e}")
        await session.rollback()
    wb.close()

    logger.info(f"Flexible import '{filename}' sheet '{target_sheet}': {imported} imported, {skipped} skipped, {errors} errors")

    return {
        "import_file_id": import_file.id,
        "filename": filename,
        "sheet": target_sheet,
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "columns_mapped": {headers[k]: v for k, v in mapping.items()},
        "columns_unmapped": unmapped_cols,
        "total_columns_detected": len([h for h in headers if h]),
    }
