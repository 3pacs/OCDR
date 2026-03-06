"""
Excel Import Engine (F-01).

Handles OCMRI.xlsx import: parses the 'Current' sheet, maps columns by header
name (case-insensitive), converts Excel serial dates, deduplicates, and
batch-inserts into billing_records.

Supports both the original 22-column layout and the updated 23-column layout
where "ID" was renamed to "Chart ID" and a new "Patient ID" column was added.
"""

import io
import logging
from datetime import date, datetime, timedelta

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.app.models.billing import BillingRecord

logger = logging.getLogger(__name__)

# Header name → BillingRecord field name (case-insensitive matching)
# Multiple aliases per field handle both old and new OCMRI layouts.
HEADER_MAP = {
    # Core fields
    "patient": "patient_name",
    "doctor": "referring_doctor",
    "scan": "scan_type",
    "gado": "gado_used",
    "insurance": "insurance_carrier",
    "type": "modality",
    "date": "service_date",
    "primary": "primary_payment",
    "secondary": "secondary_payment",
    "total": "total_payment",
    "extra": "extra_charges",
    "readby": "reading_physician",
    "read by": "reading_physician",

    # ID columns: old layout had "ID" for jacket number,
    # new layout renamed it "Chart ID" and added "Patient ID"
    "id": "patient_id",                  # old layout generic "ID" = jacket/chart number
    "jacket id": "patient_id",           # alternate name
    "chart id": "patient_id",            # new layout: renamed column
    "chart number": "patient_id",
    "patient id": "patient_id_new",      # new layout: new column (Topaz patient number)

    # Other source columns
    "birth date": "birth_date",
    "patient name": "patient_name_display",
    "s date": "schedule_date",
    "modalities": "modality_code",
    "description": "description",
    "month": "service_month",
    "year": "service_year",
    "new": "is_new_patient",

    # Topaz ID column (present in both layouts)
    "topaz id": "topaz_id",

    # Payer group (column W/X depending on layout)
    "payer group": "payer_group",
}

# Legacy fallback: positional index mapping for files with no header row
COL_MAP_LEGACY = {
    0: "patient_name",        # A - Patient
    1: "referring_doctor",     # B - Doctor
    2: "scan_type",            # C - Scan
    3: "gado_used",            # D - Gado
    4: "insurance_carrier",    # E - Insurance
    5: "modality",             # F - Type
    6: "service_date",         # G - Date
    7: "primary_payment",      # H - Primary
    8: "secondary_payment",    # I - Secondary
    9: "total_payment",        # J - Total
    10: "extra_charges",       # K - Extra
    11: "reading_physician",   # L - ReadBy
    12: "patient_id",          # M - Jacket ID / Chart ID
    13: "birth_date",          # N - Birth Date
    14: "patient_name_display",  # O - Patient Name
    15: "schedule_date",       # P - S Date
    16: "modality_code",       # Q - Modalities
    17: "description",         # R - Description
    18: "service_month",       # S - Month
    19: "service_year",        # T - Year
    20: "is_new_patient",      # U - New
    21: "topaz_id",            # V - Topaz ID
    22: "payer_group",         # W - Payer Group
}

EXCEL_EPOCH = date(1899, 12, 30)
BATCH_SIZE = 500

# Payer codes that should be normalized to 'SELF PAY'
SELFPAY_VARIANTS = {"SELFPAY", "SELF-PAY", "SELF PAY", "CASH"}


def _excel_serial_to_date(serial) -> date | None:
    """Convert Excel serial date number to Python date."""
    if serial is None:
        return None
    if isinstance(serial, datetime):
        return serial.date()
    if isinstance(serial, date):
        return serial
    try:
        serial = int(float(serial))
        if serial < 1:
            return None
        return EXCEL_EPOCH + timedelta(days=serial)
    except (ValueError, TypeError, OverflowError):
        return None


def _parse_bool(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("YES", "TRUE", "1", "Y")


def _clean_text(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip().upper()
    return s if s else None


def _parse_money(val) -> float:
    if val is None:
        return 0.0
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return 0.0


def _normalize_carrier(carrier: str) -> str:
    """Normalize insurance carrier code (BR-10)."""
    if carrier in SELFPAY_VARIANTS:
        return "SELF PAY"
    return carrier


def _derive_psma(description: str | None, modality: str | None) -> bool:
    """Detect PSMA PET scans (BR-02)."""
    if not description:
        return False
    desc_upper = description.upper()
    if "PSMA" in desc_upper:
        return True
    if modality == "PET" and ("GA-68" in desc_upper or "GALLIUM" in desc_upper):
        return True
    return False


def _safe_int_str(val) -> str | None:
    """Convert a numeric Excel value to a clean integer string (e.g., 9125.0 → '9125')."""
    if val is None:
        return None
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        s = str(val).strip()
        return s if s else None


def _date_is_reasonable(d: date | None) -> bool:
    """Check if a parsed date is within a reasonable range for billing data."""
    if d is None:
        return False
    return date(2010, 1, 1) <= d <= date(2035, 12, 31)


def _derive_service_date(month_val, year_val) -> date | None:
    """Derive service_date from separate month and year columns."""
    if month_val is None or year_val is None:
        return None
    try:
        month = int(float(month_val))
        year = int(float(year_val))
        if 1 <= month <= 12 and 2010 <= year <= 2035:
            return date(year, month, 1)
    except (ValueError, TypeError):
        pass
    return None


def _detect_headers(ws) -> tuple[dict[int, str], int]:
    """
    Detect column headers from the first row of the worksheet.

    Returns (col_index_to_field mapping, header_row_number).
    Falls back to legacy positional mapping if no headers detected.
    """
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None:
        return COL_MAP_LEGACY, 1

    # Try header-based detection
    col_map = {}
    used_fields = set()
    header_count = 0

    for col_idx, cell in enumerate(first_row):
        if cell is None:
            continue
        header = str(cell).strip().lower()
        if not header:
            continue

        # Try exact match against HEADER_MAP
        if header in HEADER_MAP:
            field = HEADER_MAP[header]
            if field not in used_fields:
                col_map[col_idx] = field
                used_fields.add(field)
                header_count += 1
                continue

        # Try partial/fuzzy match — check if any HEADER_MAP key is contained
        for alias, field in HEADER_MAP.items():
            if field in used_fields:
                continue
            if alias in header or header in alias:
                col_map[col_idx] = field
                used_fields.add(field)
                header_count += 1
                break

    # If we matched at least 5 headers, use header-based mapping
    if header_count >= 5:
        logger.info(f"Header-based detection: matched {header_count} columns")
        return col_map, 1

    # Fall back to legacy positional mapping
    logger.info("Falling back to legacy positional column mapping")
    return COL_MAP_LEGACY, 0


def _parse_row(row: tuple, col_map: dict[int, str]) -> dict | None:
    """Parse a single Excel row into a dict for BillingRecord using header mapping."""
    # Build raw dict from column mapping
    raw = {}
    extra = {}
    for col_idx, val in enumerate(row):
        if col_idx in col_map:
            raw[col_map[col_idx]] = val
        elif val is not None:
            extra[f"col_{col_idx}"] = str(val).strip()

    patient_name = _clean_text(raw.get("patient_name"))
    if not patient_name:
        return None

    referring_doctor = _clean_text(raw.get("referring_doctor"))
    scan_type = _clean_text(raw.get("scan_type"))
    insurance_carrier = _clean_text(raw.get("insurance_carrier"))
    modality = _clean_text(raw.get("modality"))

    # Parse service_date — validate it's reasonable, otherwise the "Date"
    # column may contain non-date data (e.g. patient names or IDs from
    # Candelis/Purview raw exports where columns shifted).
    service_date = _excel_serial_to_date(raw.get("service_date"))
    if not _date_is_reasonable(service_date):
        # Store the original value in extra if it was non-null
        raw_date_val = raw.get("service_date")
        if raw_date_val is not None:
            extra[f"raw_date_col"] = str(raw_date_val).strip()
        service_date = None

    # Derive service_date from month/year when direct date column is missing
    if service_date is None:
        service_date = _derive_service_date(
            raw.get("service_month"), raw.get("service_year")
        )

    # Validate schedule_date similarly — may contain non-date data
    raw_sched = raw.get("schedule_date")
    schedule_date = _excel_serial_to_date(raw_sched)
    if not _date_is_reasonable(schedule_date):
        if raw_sched is not None:
            extra[f"raw_sdate_col"] = str(raw_sched).strip()
        schedule_date = None

    # Required fields check
    if not all([referring_doctor, scan_type, insurance_carrier, modality, service_date]):
        return None

    insurance_carrier = _normalize_carrier(insurance_carrier)
    description = _clean_text(raw.get("description"))

    # Handle patient_id (chart/jacket number)
    patient_id = None
    pid_val = raw.get("patient_id")
    if pid_val is not None:
        try:
            patient_id = int(float(pid_val))
        except (ValueError, TypeError):
            pass

    # Handle topaz_id — from Topaz ID column or the new Patient ID column
    topaz_id = _safe_int_str(raw.get("topaz_id"))
    # If we have a "patient_id_new" column (new layout's "Patient ID"),
    # use it as topaz_id when the dedicated topaz column is empty
    patient_id_new = _safe_int_str(raw.get("patient_id_new"))
    if patient_id_new and not topaz_id:
        topaz_id = patient_id_new
    # Store payer_group in extra_data
    payer_group = _clean_text(raw.get("payer_group"))
    if payer_group:
        extra["payer_group"] = payer_group
    # Also preserve the new Patient ID column value in extra_data
    if patient_id_new:
        extra["patient_id_new"] = patient_id_new

    return {
        "patient_name": patient_name,
        "referring_doctor": referring_doctor,
        "scan_type": scan_type,
        "gado_used": _parse_bool(raw.get("gado_used")),
        "insurance_carrier": insurance_carrier,
        "modality": modality,
        "service_date": service_date,
        "primary_payment": _parse_money(raw.get("primary_payment")),
        "secondary_payment": _parse_money(raw.get("secondary_payment")),
        "total_payment": _parse_money(raw.get("total_payment")),
        "extra_charges": _parse_money(raw.get("extra_charges")),
        "reading_physician": _clean_text(raw.get("reading_physician")),
        "patient_id": patient_id,
        "birth_date": _excel_serial_to_date(raw.get("birth_date")),
        "patient_name_display": _clean_text(raw.get("patient_name_display")),
        "schedule_date": schedule_date,
        "modality_code": _clean_text(raw.get("modality_code")),
        "description": description,
        "service_month": _clean_text(raw.get("service_month")),
        "service_year": _clean_text(raw.get("service_year")),
        "is_new_patient": _parse_bool(raw.get("is_new_patient")),
        "topaz_id": topaz_id,
        "is_psma": _derive_psma(description, modality),
        "import_source": "EXCEL_IMPORT",
        "extra_data": extra if extra else None,
    }


async def import_excel(
    file_content: bytes,
    session: AsyncSession,
    sheet_name: str = "Current",
) -> dict:
    """
    Import an OCMRI Excel file into billing_records.

    Uses header-based column detection so the import works regardless of
    column order or count. Falls back to positional mapping for legacy
    files without headers.

    Returns dict with imported, skipped, errors counts + detected mapping.
    """
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {available}")

    ws = wb[sheet_name]

    # Detect column layout
    col_map, header_row = _detect_headers(ws)
    data_start_row = header_row + 1 if header_row > 0 else 2

    # Build readable mapping for response
    detected_mapping = {}
    for col_idx, field in sorted(col_map.items()):
        col_letter = chr(ord('A') + col_idx) if col_idx < 26 else f"col_{col_idx}"
        detected_mapping[col_letter] = field

    # Load existing dedup keys
    existing_result = await session.execute(
        select(
            BillingRecord.patient_name,
            BillingRecord.service_date,
            BillingRecord.scan_type,
            BillingRecord.modality,
        )
    )
    existing_keys = {
        (r[0], r[1], r[2], r[3]) for r in existing_result.fetchall()
    }

    imported = 0
    skipped = 0
    errors = 0
    batch: list[BillingRecord] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        try:
            parsed = _parse_row(row, col_map)
            if parsed is None:
                skipped += 1
                continue

            dedup_key = (
                parsed["patient_name"],
                parsed["service_date"],
                parsed["scan_type"],
                parsed["modality"],
            )
            if dedup_key in existing_keys:
                skipped += 1
                continue

            # Validate before inserting
            from backend.app.analytics.data_validation import validate_billing_record
            validation_errors = [v for v in validate_billing_record(parsed) if v.severity == "ERROR"]
            if validation_errors:
                logger.warning(f"Row {row_idx} validation failed: {validation_errors[0].message}")
                errors += 1
                continue

            existing_keys.add(dedup_key)
            batch.append(BillingRecord(**parsed))
            imported += 1

            if len(batch) >= BATCH_SIZE:
                session.add_all(batch)
                await session.flush()
                batch = []

        except Exception as e:
            logger.warning(f"Row {row_idx} error: {e}")
            errors += 1

    # Flush remaining batch
    if batch:
        session.add_all(batch)
        await session.flush()

    await session.commit()
    wb.close()

    logger.info(f"Excel import complete: {imported} imported, {skipped} skipped, {errors} errors")
    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "detected_mapping": detected_mapping,
        "total_columns": len(col_map),
    }
