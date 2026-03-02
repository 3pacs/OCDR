"""
Excel writer — creates NEW workbooks only.

OCMRI.xlsx is NEVER modified.  All output goes to fresh workbooks:
  - Import staging workbooks (for staff review before manual copy)
  - ERA parsed output
  - Reconciliation workbook (see ``reconciliation.py``)

Every workbook gets:
  - Auto-filters on the header row
  - Auto-width columns (up to 40 chars)
  - Frozen header row
  - Basic formatting (bold header, number formats for currency/dates)
"""

from pathlib import Path
from datetime import date
from decimal import Decimal
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

from ocdr import logger
from ocdr.normalizers import date_to_excel_serial
from ocdr.excel_reader import COLUMN_MAP, FIELD_TO_COL


# ── Styling constants ─────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4",
                           fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
CURRENCY_FMT = '#,##0.00'
DATE_FMT = 'MM/DD/YYYY'
PCT_FMT = '0%'

# Review flag colours
FLAG_FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                fill_type="solid")
FLAG_FILL_RED = PatternFill(start_color="FF6666", end_color="FF6666",
                             fill_type="solid")
FLAG_FILL_GREEN = PatternFill(start_color="90EE90", end_color="90EE90",
                               fill_type="solid")


def write_import_staging(output_path: str | Path,
                          records: list[dict],
                          sheet_name: str = "Staging") -> Path:
    """Create a staging workbook with imported data in OCMRI column format.

    Includes extra columns beyond V for:
      - suggested_modality
      - modality_confidence
      - review_flags
      - is_guess

    Staff reviews this sheet and manually copies confirmed rows into OCMRI.xlsx.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── Header row ──
    # Standard OCMRI columns (A-V) + review columns
    ocmri_headers = [COLUMN_MAP[i] for i in range(1, 23)]
    extra_headers = [
        "suggested_modality", "modality_confidence", "review_flags",
        "is_guess", "ae_title", "description", "modality_code",
    ]
    all_headers = ocmri_headers + extra_headers

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ──
    for row_idx, record in enumerate(records, start=2):
        # Standard OCMRI columns
        for col_num in range(1, 23):
            field = COLUMN_MAP[col_num]
            value = _format_cell_value(record.get(field), field)
            cell = ws.cell(row=row_idx, column=col_num, value=value)
            _apply_format(cell, field)

        # Extra review columns
        extra_start = 23
        ws.cell(row=row_idx, column=extra_start,
                value=record.get("suggested_modality", ""))
        conf = record.get("modality_confidence", "")
        conf_cell = ws.cell(row=row_idx, column=extra_start + 1, value=conf)
        if isinstance(conf, (int, float)) and conf < 1.0:
            conf_cell.fill = FLAG_FILL_YELLOW

        flags = record.get("review_flags", [])
        flags_cell = ws.cell(row=row_idx, column=extra_start + 2,
                              value="; ".join(flags) if flags else "")
        if flags:
            flags_cell.fill = FLAG_FILL_YELLOW

        ws.cell(row=row_idx, column=extra_start + 3,
                value="YES" if record.get("is_guess") else "")
        ws.cell(row=row_idx, column=extra_start + 4,
                value=record.get("ae_title", ""))
        ws.cell(row=row_idx, column=extra_start + 5,
                value=record.get("description", ""))
        ws.cell(row=row_idx, column=extra_start + 6,
                value=record.get("modality_code", ""))

    _finalize_sheet(ws, len(all_headers))
    wb.save(str(path))
    wb.close()

    logger.log_decision(
        "excel_write_staging",
        {"output": str(path)},
        {"rows": len(records), "columns": len(all_headers)},
        flags=[],
        reasoning=f"Wrote {len(records)} records to staging workbook.",
    )
    return path


def write_era_output(output_path: str | Path,
                      parsed_files: list[dict]) -> Path:
    """Create a workbook with parsed 835 payment/claim data.

    Two sheets:
      - "Payments": One row per 835 file (payment summary)
      - "Claims": One row per claim line (detail)
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ── Payments sheet ──
    ws_pay = wb.active
    ws_pay.title = "Payments"
    pay_headers = [
        "file", "payer_name", "check_eft_number", "payment_method",
        "payment_amount", "payment_date", "claim_count",
    ]
    _write_header(ws_pay, pay_headers)

    for row_idx, pf in enumerate(parsed_files, start=2):
        ws_pay.cell(row=row_idx, column=1, value=pf.get("file", ""))
        ws_pay.cell(row=row_idx, column=2, value=pf.get("payer_name", ""))
        ws_pay.cell(row=row_idx, column=3, value=pf.get("check_eft_number", ""))
        ws_pay.cell(row=row_idx, column=4,
                    value=pf.get("payment", {}).get("method", ""))
        amt_cell = ws_pay.cell(
            row=row_idx, column=5,
            value=float(pf.get("payment", {}).get("amount", 0)))
        amt_cell.number_format = CURRENCY_FMT
        pay_date = pf.get("payment", {}).get("date")
        ws_pay.cell(row=row_idx, column=6,
                    value=pay_date.strftime("%m/%d/%Y") if pay_date else "")
        ws_pay.cell(row=row_idx, column=7,
                    value=len(pf.get("claims", [])))

    _finalize_sheet(ws_pay, len(pay_headers))

    # ── Claims sheet ──
    ws_claims = wb.create_sheet("Claims")
    claim_headers = [
        "source_file", "payer_name", "check_number", "claim_id",
        "claim_status", "patient_name", "service_date",
        "billed_amount", "paid_amount", "patient_responsibility",
        "cpt_codes", "adjustment_summary", "rendering_provider",
    ]
    _write_header(ws_claims, claim_headers)

    row_idx = 2
    for pf in parsed_files:
        for claim in pf.get("claims", []):
            ws_claims.cell(row=row_idx, column=1, value=pf.get("file", ""))
            ws_claims.cell(row=row_idx, column=2, value=pf.get("payer_name", ""))
            ws_claims.cell(row=row_idx, column=3,
                           value=pf.get("check_eft_number", ""))
            ws_claims.cell(row=row_idx, column=4,
                           value=claim.get("claim_id", ""))
            ws_claims.cell(row=row_idx, column=5,
                           value=claim.get("claim_status", ""))
            ws_claims.cell(row=row_idx, column=6,
                           value=claim.get("patient_name", ""))
            sd = claim.get("service_date")
            ws_claims.cell(row=row_idx, column=7,
                           value=sd.strftime("%m/%d/%Y") if sd else "")
            billed_cell = ws_claims.cell(
                row=row_idx, column=8,
                value=float(claim.get("billed_amount", 0)))
            billed_cell.number_format = CURRENCY_FMT
            paid_cell = ws_claims.cell(
                row=row_idx, column=9,
                value=float(claim.get("paid_amount", 0)))
            paid_cell.number_format = CURRENCY_FMT
            pr_cell = ws_claims.cell(
                row=row_idx, column=10,
                value=float(claim.get("patient_responsibility", 0)))
            pr_cell.number_format = CURRENCY_FMT

            # CPT codes from service lines
            cpts = [sl.get("cpt_code", "")
                    for sl in claim.get("service_lines", [])]
            ws_claims.cell(row=row_idx, column=11,
                           value=", ".join(cpts) if cpts else "")

            # Adjustment summary
            adjs = claim.get("adjustments", [])
            adj_parts = []
            for a in adjs:
                for sub in a.get("adjustments", []):
                    adj_parts.append(
                        f"{a.get('group_code', '')}:{sub.get('reason_code', '')}="
                        f"${sub.get('amount', 0)}"
                    )
            ws_claims.cell(row=row_idx, column=12,
                           value="; ".join(adj_parts) if adj_parts else "")
            ws_claims.cell(row=row_idx, column=13,
                           value=claim.get("rendering_provider", ""))
            row_idx += 1

    _finalize_sheet(ws_claims, len(claim_headers))

    wb.save(str(path))
    wb.close()

    logger.log_decision(
        "excel_write_era",
        {"output": str(path)},
        {"files": len(parsed_files),
         "total_claims": sum(len(pf.get("claims", [])) for pf in parsed_files)},
        flags=[],
    )
    return path


def write_generic_sheet(wb: openpyxl.Workbook,
                         sheet_name: str,
                         headers: list[str],
                         rows: list[list[Any]],
                         currency_cols: Optional[list[int]] = None,
                         date_cols: Optional[list[int]] = None,
                         pct_cols: Optional[list[int]] = None,
                         highlight_condition: Optional[callable] = None) -> None:
    """Write a formatted sheet into an existing workbook.

    Args:
        wb: Target workbook.
        sheet_name: Name for the new sheet.
        headers: Column header strings.
        rows: List of row-value lists.
        currency_cols: 0-indexed column indices to format as currency.
        date_cols: 0-indexed column indices to format as dates.
        pct_cols: 0-indexed column indices to format as percentages.
        highlight_condition: Optional callable(row_data) → fill color name
            ("yellow", "red", "green") or None.
    """
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, headers)

    fill_map = {
        "yellow": FLAG_FILL_YELLOW,
        "red": FLAG_FILL_RED,
        "green": FLAG_FILL_GREEN,
    }

    for row_idx, row_data in enumerate(rows, start=2):
        highlight = None
        if highlight_condition:
            highlight = highlight_condition(row_data)

        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx, column=col_idx + 1,
                           value=_to_writable(value))

            if currency_cols and col_idx in currency_cols:
                cell.number_format = CURRENCY_FMT
            if date_cols and col_idx in date_cols:
                cell.number_format = DATE_FMT
            if pct_cols and col_idx in pct_cols:
                cell.number_format = PCT_FMT

            if highlight and highlight in fill_map:
                cell.fill = fill_map[highlight]

    _finalize_sheet(ws, len(headers))


# ── Helpers ───────────────────────────────────────────────────────────────

def _write_header(ws, headers: list[str]) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _finalize_sheet(ws, num_cols: int) -> None:
    """Apply auto-filters, freeze panes, and auto-width to a sheet."""
    # Auto-filter
    if ws.max_row and ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{ws.max_row}"

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-width columns (cap at 40)
    for col_idx in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                 values_only=True):
            for cell_val in row:
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _format_cell_value(value: Any, field: str) -> Any:
    """Convert a record value to an Excel-friendly format."""
    if value is None:
        return ""
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bool):
        return "YES" if value else ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    return value


def _apply_format(cell, field: str) -> None:
    """Apply number format based on field type."""
    if field in ("primary_payment", "secondary_payment",
                 "total_payment", "extra_charges"):
        cell.number_format = CURRENCY_FMT
    elif field in ("service_date", "birth_date", "schedule_date"):
        # Already formatted as string; no special format needed
        pass


def write_payment_applied(output_path: str | Path,
                           approved_records: list[dict]) -> Path:
    """Write approved payments in OCMRI A-V format + payment metadata columns.

    Staff copies confirmed rows from this workbook into OCMRI.xlsx.
    Columns M/N/O (primary/secondary/total) are filled from the 835 data.
    Extra columns beyond V provide 835 audit trail info.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approved Payments"

    # Standard OCMRI columns (A-V) + payment audit columns
    ocmri_headers = [COLUMN_MAP[i] for i in range(1, 23)]
    extra_headers = [
        "claim_id", "check_eft_number", "payer_835", "payment_date",
        "payment_method", "cpt_codes", "claim_billed", "claim_status",
        "match_score", "approval_timestamp",
    ]
    all_headers = ocmri_headers + extra_headers

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(approved_records, start=2):
        # Standard OCMRI columns
        for col_num in range(1, 23):
            field = COLUMN_MAP[col_num]
            value = _format_cell_value(record.get(field), field)
            cell = ws.cell(row=row_idx, column=col_num, value=value)
            _apply_format(cell, field)

        # Payment audit columns
        extra_start = 23
        ws.cell(row=row_idx, column=extra_start,
                value=record.get("claim_id", ""))
        ws.cell(row=row_idx, column=extra_start + 1,
                value=record.get("check_eft_number", ""))
        ws.cell(row=row_idx, column=extra_start + 2,
                value=record.get("payer_835", ""))
        pd = record.get("payment_date")
        ws.cell(row=row_idx, column=extra_start + 3,
                value=pd.strftime("%m/%d/%Y") if pd else "")
        ws.cell(row=row_idx, column=extra_start + 4,
                value=record.get("payment_method", ""))
        ws.cell(row=row_idx, column=extra_start + 5,
                value=record.get("cpt_codes", ""))
        billed_cell = ws.cell(row=row_idx, column=extra_start + 6,
                               value=float(record.get("claim_billed", 0) or 0))
        billed_cell.number_format = CURRENCY_FMT
        ws.cell(row=row_idx, column=extra_start + 7,
                value=record.get("claim_status", ""))
        score_cell = ws.cell(row=row_idx, column=extra_start + 8,
                              value=record.get("match_score", 0))
        score_cell.number_format = '0.00'
        ws.cell(row=row_idx, column=extra_start + 9,
                value=record.get("approval_timestamp", ""))

    _finalize_sheet(ws, len(all_headers))
    wb.save(str(path))
    wb.close()

    logger.log_decision(
        "excel_write_payment_applied",
        {"output": str(path)},
        {"rows": len(approved_records)},
        flags=[],
        reasoning=f"Wrote {len(approved_records)} approved payment records.",
    )
    return path


def _to_writable(value: Any) -> Any:
    """Convert any value to something openpyxl can write."""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, bool):
        return "YES" if value else ""
    if isinstance(value, (list, tuple)):
        return "; ".join(str(v) for v in value)
    return value
