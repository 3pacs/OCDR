"""
Excel reader for OCMRI.xlsx (READ-ONLY).

Reads the existing billing workbook and converts rows into dicts
for use by the reconciliation engine and other analysis tools.

OCMRI.xlsx is NEVER modified by this system.  All output goes to
new workbooks created by ``excel_writer.py``.

Column layout (from BUILD_SPEC):
  A = patient_name_display     L = reading_physician
  B = service_date (Excel serial)  M = primary_payment
  C = scan_type                N = secondary_payment
  D = modality                 O = total_payment
  E = gado_used (YES/NO)       P = extra_charges
  F = is_new_patient (NEW/blank)  Q = service_month
  G = referring_doctor          R = service_year
  H = insurance_carrier         S = patient_id
  I = patient_name             T = is_research
  J = birth_date               U = source
  K = schedule_date            V = notes
"""

from pathlib import Path
from decimal import Decimal
from typing import Optional

import openpyxl

from ocdr import logger
from ocdr.normalizers import (
    normalize_patient_name, parse_date_flexible, normalize_decimal,
    normalize_modality, normalize_scan_type, normalize_gado,
    normalize_payer_code, normalize_text, derive_month, derive_year,
)


# Column letter → field name (1-indexed column numbers)
COLUMN_MAP = {
    1:  "patient_name_display",
    2:  "service_date",
    3:  "scan_type",
    4:  "modality",
    5:  "gado_used",
    6:  "is_new_patient",
    7:  "referring_doctor",
    8:  "insurance_carrier",
    9:  "patient_name",
    10: "birth_date",
    11: "schedule_date",
    12: "reading_physician",
    13: "primary_payment",
    14: "secondary_payment",
    15: "total_payment",
    16: "extra_charges",
    17: "service_month",
    18: "service_year",
    19: "patient_id",
    20: "is_research",
    21: "source",
    22: "notes",
}

# Reverse: field name → column number
FIELD_TO_COL = {v: k for k, v in COLUMN_MAP.items()}


def read_ocmri(filepath: str | Path,
               sheet_name: str = "Current") -> list[dict]:
    """Read all data rows from an OCMRI workbook.

    Returns a list of dicts with normalized field values.
    Row 1 is assumed to be the header and is skipped.
    """
    path = Path(filepath)
    if not path.exists():
        logger.log_warning("excel_read", f"File not found: {path}", {})
        return []

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        logger.log_warning(
            "excel_read",
            f"Sheet '{sheet_name}' not found in {path.name}",
            {"available_sheets": wb.sheetnames},
        )
        wb.close()
        return []

    ws = wb[sheet_name]
    records: list[dict] = []
    errors: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue  # Skip empty rows

        try:
            record = _parse_excel_row(row, row_idx)
            if record is not None:
                records.append(record)
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})
            logger.log_error(
                "excel_read_row",
                {"row": row_idx, "values": [str(v)[:50] for v in row[:10]]},
                e,
                suggested_fix="Check that OCMRI.xlsx column layout matches "
                              "the expected A-V schema from BUILD_SPEC.",
            )

    wb.close()

    logger.log_import_summary(
        source=f"ocmri:{path.name}:{sheet_name}",
        records_in=row_idx - 1 if 'row_idx' in dir() else 0,
        records_out=len(records),
        errors=errors,
        warnings=[],
    )
    return records


def _parse_excel_row(row: tuple, row_idx: int) -> Optional[dict]:
    """Parse a single Excel row into a normalized billing record dict."""

    def cell(col_num: int):
        """Get cell value by 1-indexed column number."""
        idx = col_num - 1
        return row[idx] if idx < len(row) else None

    # Patient name is the key field — skip rows without it
    raw_name = cell(FIELD_TO_COL["patient_name"])
    display_name = cell(FIELD_TO_COL["patient_name_display"])
    if raw_name is None and display_name is None:
        return None

    patient_name = normalize_patient_name(raw_name or display_name)
    if not patient_name:
        return None

    # Parse dates
    service_date = parse_date_flexible(cell(FIELD_TO_COL["service_date"]))
    birth_date = parse_date_flexible(cell(FIELD_TO_COL["birth_date"]))
    schedule_date = parse_date_flexible(cell(FIELD_TO_COL["schedule_date"]))

    # Parse modality and scan type
    modality = normalize_modality(cell(FIELD_TO_COL["modality"]))
    scan_type = normalize_scan_type(cell(FIELD_TO_COL["scan_type"]))

    # Parse monetary values
    primary = normalize_decimal(cell(FIELD_TO_COL["primary_payment"]))
    secondary = normalize_decimal(cell(FIELD_TO_COL["secondary_payment"]))
    total = normalize_decimal(cell(FIELD_TO_COL["total_payment"]))
    extra = normalize_decimal(cell(FIELD_TO_COL["extra_charges"]))

    # Parse boolean/flag fields
    gado_raw = cell(FIELD_TO_COL["gado_used"])
    gado = normalize_gado(gado_raw)
    new_patient_raw = cell(FIELD_TO_COL["is_new_patient"])
    is_new = str(new_patient_raw or "").strip().upper() == "NEW"

    # Text fields
    referring = normalize_text(cell(FIELD_TO_COL["referring_doctor"]))
    carrier = normalize_payer_code(cell(FIELD_TO_COL["insurance_carrier"]))
    reading = normalize_text(cell(FIELD_TO_COL["reading_physician"]))
    source = normalize_text(cell(FIELD_TO_COL["source"]))
    notes = str(cell(FIELD_TO_COL["notes"]) or "").strip()

    # Patient ID
    pid_raw = cell(FIELD_TO_COL["patient_id"])
    patient_id = int(pid_raw) if pid_raw and str(pid_raw).strip().isdigit() else None

    # Research flag
    research_raw = cell(FIELD_TO_COL["is_research"])
    is_research = str(research_raw or "").strip().upper() in ("TRUE", "YES", "1", "RESEARCH")

    return {
        "patient_name": patient_name,
        "patient_name_display": normalize_text(display_name) or patient_name,
        "patient_id": patient_id,
        "birth_date": birth_date,
        "service_date": service_date,
        "schedule_date": schedule_date,
        "scan_type": scan_type,
        "modality": modality,
        "gado_used": gado,
        "is_new_patient": is_new,
        "referring_doctor": referring,
        "insurance_carrier": carrier,
        "reading_physician": reading,
        "primary_payment": primary,
        "secondary_payment": secondary,
        "total_payment": total,
        "extra_charges": extra,
        "service_month": derive_month(service_date),
        "service_year": derive_year(service_date),
        "is_research": is_research,
        "source": source,
        "notes": notes,
        "source_row": row_idx,
        "review_flags": [],
    }


def get_last_row(filepath: str | Path,
                 sheet_name: str = "Current") -> int:
    """Find the last populated row number in a sheet.

    Useful for knowing where new data would be appended.
    """
    path = Path(filepath)
    if not path.exists():
        return 1

    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb[sheet_name]
    last = ws.max_row or 1
    wb.close()
    return last


def get_sheet_names(filepath: str | Path) -> list[str]:
    """Return all sheet names in a workbook."""
    path = Path(filepath)
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True)
    names = wb.sheetnames
    wb.close()
    return names
