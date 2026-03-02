"""
Reconciliation workbook generator.

Produces ``OCDR_Reconciliation.xlsx`` with analysis sheets:

  1. **Billing Records** — mirror of OCMRI.xlsx + computed columns
  2. **ERA Payments** — summary of 835 payment files
  3. **ERA Claims** — detail of each 835 claim
  4. **Matched Claims** — successfully matched billing ↔ ERA pairs
  5. **Unmatched Claims** — 835 claims with no billing match
  6. **Underpayments** — claims paid < 80% of expected rate
  7. **Denials** — claims with $0 payment, sorted by recoverability
  8. **Filing Deadlines** — approaching or past filing deadlines
  9. **Missing Secondary** — M/M and CALOPTIMA claims without secondary

All business rules are applied during generation.  Every sheet gets
auto-filters, frozen headers, and conditional formatting.
"""

from pathlib import Path
from datetime import date
from decimal import Decimal
from typing import Optional

import openpyxl

from ocdr import logger
from ocdr.business_rules import (
    detect_cap_exceptions, detect_underpayments, detect_denials,
    detect_filing_issues, detect_missing_secondary,
    recoverability_score, detect_duplicates,
)
from ocdr.config import get_expected_rate, MATCH_AUTO_ACCEPT
from ocdr.excel_writer import (
    write_generic_sheet, FLAG_FILL_YELLOW, FLAG_FILL_RED, FLAG_FILL_GREEN,
    _finalize_sheet, _write_header, _to_writable, CURRENCY_FMT, PCT_FMT,
)


def generate_reconciliation(billing_records: list[dict],
                             era_data: list[dict],
                             match_results: Optional[list[dict]] = None,
                             output_path: str | Path = None,
                             as_of: date | None = None) -> Path:
    """Generate the full reconciliation workbook.

    Args:
        billing_records: Records read from OCMRI.xlsx via ``excel_reader``.
        era_data: Parsed 835 files (list of dicts from ``era_835_parser``).
        match_results: Pre-computed match results (from ``match_835_to_billing``).
            If None, matching is skipped and only billing analysis sheets are created.
        output_path: Where to save the workbook. Defaults to config path.
        as_of: Reference date for filing deadline calculations.

    Returns:
        Path to the generated workbook.
    """
    from ocdr.config import RECONCILIATION_PATH
    if output_path is None:
        output_path = RECONCILIATION_PATH
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if as_of is None:
        as_of = date.today()

    wb = openpyxl.Workbook()
    # Remove default sheet (we create our own)
    wb.remove(wb.active)

    # ── Enrich billing records with computed fields ──
    billing_records = _enrich_billing(billing_records, as_of)

    # ── Sheet 1: Billing Records ──
    _write_billing_sheet(wb, billing_records)

    # ── Sheet 2 & 3: ERA Payments and Claims ──
    if era_data:
        _write_era_payments_sheet(wb, era_data)
        _write_era_claims_sheet(wb, era_data)

    # ── Sheet 4 & 5: Matched and Unmatched Claims ──
    if match_results:
        _write_matched_sheet(wb, match_results)
        _write_unmatched_sheet(wb, match_results)

    # ── Sheet 6: Underpayments ──
    _write_underpayments_sheet(wb, billing_records)

    # ── Sheet 7: Denials ──
    _write_denials_sheet(wb, billing_records, as_of)

    # ── Sheet 8: Filing Deadlines ──
    _write_filing_sheet(wb, billing_records, as_of)

    # ── Sheet 9: Missing Secondary ──
    _write_missing_secondary_sheet(wb, billing_records)

    # ── Sheet 10: Duplicates ──
    _write_duplicates_sheet(wb, billing_records)

    wb.save(str(path))
    wb.close()

    logger.log_decision(
        "reconciliation_generate",
        {"billing_count": len(billing_records),
         "era_files": len(era_data) if era_data else 0,
         "match_results": len(match_results) if match_results else 0},
        {"output": str(path), "sheets": wb.sheetnames if hasattr(wb, 'sheetnames') else []},
        flags=[],
        reasoning="Generated reconciliation workbook with all analysis sheets.",
    )
    return path


# ── Internal sheet builders ───────────────────────────────────────────────

def _enrich_billing(records: list[dict], as_of: date) -> list[dict]:
    """Add computed fields to billing records for reconciliation."""
    records = detect_cap_exceptions(records)

    for r in records:
        # Expected rate
        expected = get_expected_rate(
            r.get("modality", ""),
            r.get("insurance_carrier", "DEFAULT"),
            r.get("is_psma", False),
            r.get("gado_used", False),
        )
        r["expected_rate"] = expected

        # Variance
        total = r.get("total_payment", Decimal("0"))
        if expected > 0 and total > 0:
            r["variance"] = total - expected
            r["pct_of_expected"] = float(total / expected)
        else:
            r["variance"] = Decimal("0")
            r["pct_of_expected"] = 0.0

    # Filing deadlines
    detect_filing_issues(records, as_of)

    return records


def _write_billing_sheet(wb: openpyxl.Workbook,
                          records: list[dict]) -> None:
    """Sheet 1: Full billing records with computed columns."""
    headers = [
        "patient_name", "service_date", "scan_type", "modality",
        "gado_used", "is_new_patient", "referring_doctor",
        "insurance_carrier", "primary_payment", "secondary_payment",
        "total_payment", "extra_charges", "expected_rate", "variance",
        "pct_of_expected", "filing_deadline", "filing_status",
        "filing_days_remaining", "is_research", "source",
        "review_flags",
    ]

    rows = []
    for r in records:
        rows.append([
            r.get("patient_name", ""),
            r.get("service_date"),
            r.get("scan_type", ""),
            r.get("modality", ""),
            "YES" if r.get("gado_used") else "",
            "NEW" if r.get("is_new_patient") else "",
            r.get("referring_doctor", ""),
            r.get("insurance_carrier", ""),
            r.get("primary_payment", Decimal("0")),
            r.get("secondary_payment", Decimal("0")),
            r.get("total_payment", Decimal("0")),
            r.get("extra_charges", Decimal("0")),
            r.get("expected_rate", Decimal("0")),
            r.get("variance", Decimal("0")),
            r.get("pct_of_expected", 0.0),
            r.get("filing_deadline"),
            r.get("filing_status", ""),
            r.get("filing_days_remaining", ""),
            "YES" if r.get("is_research") else "",
            r.get("source", ""),
            "; ".join(r.get("review_flags", [])),
        ])

    def highlight(row_data):
        status = row_data[16] if len(row_data) > 16 else ""
        if status == "PAST_DEADLINE":
            return "red"
        if status == "WARNING_30DAY":
            return "yellow"
        return None

    write_generic_sheet(
        wb, "Billing Records", headers, rows,
        currency_cols=[8, 9, 10, 11, 12, 13],
        date_cols=[1, 15],
        pct_cols=[14],
        highlight_condition=highlight,
    )


def _write_era_payments_sheet(wb: openpyxl.Workbook,
                                era_data: list[dict]) -> None:
    """Sheet 2: ERA payment summaries."""
    headers = [
        "file", "payer_name", "check_eft_number", "payment_method",
        "payment_amount", "payment_date", "claim_count",
    ]
    rows = []
    for pf in era_data:
        rows.append([
            pf.get("file", ""),
            pf.get("payer_name", ""),
            pf.get("check_eft_number", ""),
            pf.get("payment", {}).get("method", ""),
            pf.get("payment", {}).get("amount", Decimal("0")),
            pf.get("payment", {}).get("date"),
            len(pf.get("claims", [])),
        ])
    write_generic_sheet(wb, "ERA Payments", headers, rows,
                         currency_cols=[4], date_cols=[5])


def _write_era_claims_sheet(wb: openpyxl.Workbook,
                              era_data: list[dict]) -> None:
    """Sheet 3: ERA claim detail."""
    headers = [
        "source_file", "payer", "check_number", "claim_id",
        "status", "patient_name", "service_date",
        "billed_amount", "paid_amount", "adjustments",
    ]
    rows = []
    for pf in era_data:
        for claim in pf.get("claims", []):
            adj_parts = []
            for a in claim.get("adjustments", []):
                for sub in a.get("adjustments", []):
                    adj_parts.append(
                        f"{a.get('group_code', '')}:{sub.get('reason_code', '')}"
                        f"=${sub.get('amount', 0)}"
                    )
            rows.append([
                pf.get("file", ""),
                pf.get("payer_name", ""),
                pf.get("check_eft_number", ""),
                claim.get("claim_id", ""),
                claim.get("claim_status", ""),
                claim.get("patient_name", ""),
                claim.get("service_date"),
                claim.get("billed_amount", Decimal("0")),
                claim.get("paid_amount", Decimal("0")),
                "; ".join(adj_parts),
            ])
    write_generic_sheet(wb, "ERA Claims", headers, rows,
                         currency_cols=[7, 8], date_cols=[6])


def _write_matched_sheet(wb: openpyxl.Workbook,
                           match_results: list[dict]) -> None:
    """Sheet 4: Successfully matched billing ↔ ERA pairs."""
    matched = [m for m in match_results if m["status"] != "UNMATCHED"]
    headers = [
        "status", "match_score", "patient_name_billing",
        "patient_name_era", "service_date", "modality",
        "billed_amount", "paid_amount", "name_sim",
        "date_match", "modality_match", "body_part_match",
        "mismatches",
    ]
    rows = []
    for m in matched:
        br = m.get("billing_record") or {}
        cl = m.get("claim") or {}
        sc = m.get("match_score") or {}
        rows.append([
            m.get("status", ""),
            sc.get("score", 0.0),
            br.get("patient_name", ""),
            cl.get("patient_name", ""),
            br.get("service_date") or cl.get("service_date"),
            br.get("modality", ""),
            cl.get("billed_amount", Decimal("0")),
            cl.get("paid_amount", Decimal("0")),
            sc.get("name_sim", 0.0),
            sc.get("date_match", 0.0),
            sc.get("modality_match", 0.0),
            sc.get("body_part_match", 0.0),
            "; ".join(sc.get("mismatches", [])),
        ])

    def highlight(row_data):
        status = row_data[0] if row_data else ""
        if status == "AUTO_ACCEPT":
            return "green"
        if status == "REVIEW":
            return "yellow"
        return None

    write_generic_sheet(wb, "Matched Claims", headers, rows,
                         currency_cols=[6, 7], pct_cols=[1, 8, 9, 10, 11],
                         highlight_condition=highlight)


def _write_unmatched_sheet(wb: openpyxl.Workbook,
                             match_results: list[dict]) -> None:
    """Sheet 5: Unmatched 835 claims."""
    unmatched = [m for m in match_results if m["status"] == "UNMATCHED"]
    headers = [
        "claim_id", "patient_name", "service_date",
        "billed_amount", "paid_amount", "best_score",
        "mismatches",
    ]
    rows = []
    for m in unmatched:
        cl = m.get("claim") or {}
        sc = m.get("match_score") or {}
        rows.append([
            cl.get("claim_id", ""),
            cl.get("patient_name", ""),
            cl.get("service_date"),
            cl.get("billed_amount", Decimal("0")),
            cl.get("paid_amount", Decimal("0")),
            sc.get("score", 0.0),
            "; ".join(sc.get("mismatches", [])),
        ])
    write_generic_sheet(wb, "Unmatched Claims", headers, rows,
                         currency_cols=[3, 4], date_cols=[2],
                         highlight_condition=lambda r: "red")


def _write_underpayments_sheet(wb: openpyxl.Workbook,
                                 records: list[dict]) -> None:
    """Sheet 6: Underpaid claims."""
    flagged = detect_underpayments(records)
    headers = [
        "patient_name", "service_date", "modality",
        "insurance_carrier", "total_payment", "expected_rate",
        "variance", "pct_of_expected",
    ]
    rows = []
    for r in flagged:
        rows.append([
            r.get("patient_name", ""),
            r.get("service_date"),
            r.get("modality", ""),
            r.get("insurance_carrier", ""),
            r.get("total_payment", Decimal("0")),
            r.get("expected_rate", Decimal("0")),
            r.get("variance", Decimal("0")),
            r.get("pct_of_expected", 0.0),
        ])
    write_generic_sheet(wb, "Underpayments", headers, rows,
                         currency_cols=[4, 5, 6], date_cols=[1],
                         pct_cols=[7],
                         highlight_condition=lambda r: "red")


def _write_denials_sheet(wb: openpyxl.Workbook,
                          records: list[dict],
                          as_of: date) -> None:
    """Sheet 7: Denied claims sorted by recoverability."""
    denied = detect_denials(records)
    # Compute recoverability and sort
    for r in denied:
        sd = r.get("service_date")
        billed = r.get("expected_rate", Decimal("0"))
        r["recoverability"] = (
            recoverability_score(billed, sd, as_of) if sd and billed else 0.0
        )
    denied.sort(key=lambda r: r.get("recoverability", 0), reverse=True)

    headers = [
        "patient_name", "service_date", "modality",
        "insurance_carrier", "expected_rate", "recoverability_score",
        "filing_status", "filing_days_remaining",
    ]
    rows = []
    for r in denied:
        rows.append([
            r.get("patient_name", ""),
            r.get("service_date"),
            r.get("modality", ""),
            r.get("insurance_carrier", ""),
            r.get("expected_rate", Decimal("0")),
            round(r.get("recoverability", 0.0), 2),
            r.get("filing_status", ""),
            r.get("filing_days_remaining", ""),
        ])

    def highlight(row_data):
        status = row_data[6] if len(row_data) > 6 else ""
        if status == "PAST_DEADLINE":
            return "red"
        if status == "WARNING_30DAY":
            return "yellow"
        return None

    write_generic_sheet(wb, "Denials", headers, rows,
                         currency_cols=[4], date_cols=[1],
                         highlight_condition=highlight)


def _write_filing_sheet(wb: openpyxl.Workbook,
                         records: list[dict],
                         as_of: date) -> None:
    """Sheet 8: Filing deadline status for all unpaid claims."""
    past, warning = detect_filing_issues(records, as_of)
    combined = past + warning
    # Also include records with known deadlines that are safe
    safe = [r for r in records
            if r.get("filing_status") == "SAFE"
            and r.get("total_payment", Decimal("0")) == 0]
    combined.extend(safe)

    # Sort: past deadline first, then by days remaining
    combined.sort(key=lambda r: r.get("filing_days_remaining", 9999))

    headers = [
        "patient_name", "service_date", "modality",
        "insurance_carrier", "filing_deadline", "days_remaining",
        "filing_status",
    ]
    rows = []
    for r in combined:
        rows.append([
            r.get("patient_name", ""),
            r.get("service_date"),
            r.get("modality", ""),
            r.get("insurance_carrier", ""),
            r.get("filing_deadline"),
            r.get("filing_days_remaining", ""),
            r.get("filing_status", ""),
        ])

    def highlight(row_data):
        status = row_data[6] if len(row_data) > 6 else ""
        if status == "PAST_DEADLINE":
            return "red"
        if status == "WARNING_30DAY":
            return "yellow"
        if status == "SAFE":
            return "green"
        return None

    write_generic_sheet(wb, "Filing Deadlines", headers, rows,
                         date_cols=[1, 4], highlight_condition=highlight)


def _write_missing_secondary_sheet(wb: openpyxl.Workbook,
                                     records: list[dict]) -> None:
    """Sheet 9: Claims expecting secondary payment but none received."""
    missing = detect_missing_secondary(records)
    headers = [
        "patient_name", "service_date", "modality",
        "insurance_carrier", "primary_payment", "secondary_payment",
        "total_payment",
    ]
    rows = []
    for r in missing:
        rows.append([
            r.get("patient_name", ""),
            r.get("service_date"),
            r.get("modality", ""),
            r.get("insurance_carrier", ""),
            r.get("primary_payment", Decimal("0")),
            r.get("secondary_payment", Decimal("0")),
            r.get("total_payment", Decimal("0")),
        ])
    write_generic_sheet(wb, "Missing Secondary", headers, rows,
                         currency_cols=[4, 5, 6], date_cols=[1],
                         highlight_condition=lambda r: "yellow")


def _write_duplicates_sheet(wb: openpyxl.Workbook,
                              records: list[dict]) -> None:
    """Sheet 10: Potential duplicate billing records."""
    dup_groups = detect_duplicates(records)
    headers = [
        "group", "patient_name", "service_date", "scan_type",
        "modality", "insurance_carrier", "total_payment", "source",
    ]
    rows = []
    for group_idx, group in enumerate(dup_groups, start=1):
        for r in group:
            rows.append([
                f"Group {group_idx}",
                r.get("patient_name", ""),
                r.get("service_date"),
                r.get("scan_type", ""),
                r.get("modality", ""),
                r.get("insurance_carrier", ""),
                r.get("total_payment", Decimal("0")),
                r.get("source", ""),
            ])
    write_generic_sheet(wb, "Duplicates", headers, rows,
                         currency_cols=[6], date_cols=[2],
                         highlight_condition=lambda r: "yellow")
