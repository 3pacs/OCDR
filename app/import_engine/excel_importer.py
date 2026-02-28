"""Excel Import Engine for billing records (F-01).

Imports billing data from Excel workbooks (.xlsx) into billing_records table.
Supports the OCMRI.xlsx format with column mapping, date conversion,
deduplication, and batch inserts.
"""

import os

import openpyxl

from app.models import db, BillingRecord
from app.import_engine.validation import (
    parse_date, parse_float, parse_bool, normalize_modality,
    normalize_carrier, detect_psma, compute_total_payment,
    build_dedup_set, is_duplicate,
)
from app.import_engine.column_learner import enhance_column_map
from app.import_engine.normalization_learner import (
    enhanced_normalize_modality, enhanced_normalize_carrier,
)


# Column mapping: Excel header → model field (with aliases)
COLUMN_MAP = {
    "patient": "patient_name",
    "patient name": "patient_name",
    "name": "patient_name",
    "doctor": "referring_doctor",
    "referring doctor": "referring_doctor",
    "ref doctor": "referring_doctor",
    "referring": "referring_doctor",
    "scan": "scan_type",
    "scan type": "scan_type",
    "procedure": "scan_type",
    "gado": "gado_used",
    "contrast": "gado_used",
    "gadolinium": "gado_used",
    "insurance": "insurance_carrier",
    "insurance carrier": "insurance_carrier",
    "carrier": "insurance_carrier",
    "payer": "insurance_carrier",
    "type": "modality",
    "modality": "modality",
    "modalities": "modality",
    "date": "service_date",
    "service date": "service_date",
    "s date": "service_date",
    "dos": "service_date",
    "primary": "primary_payment",
    "primary payment": "primary_payment",
    "secondary": "secondary_payment",
    "secondary payment": "secondary_payment",
    "total": "total_payment",
    "total payment": "total_payment",
    "extra": "extra_charges",
    "extra charges": "extra_charges",
    "read by": "reading_physician",
    "reading physician": "reading_physician",
    "readby": "reading_physician",
    "id": "patient_id",
    "patient id": "patient_id",
    "description": "description",
    "desc": "description",
}


def _normalize_header(header):
    """Normalize a header string for column mapping."""
    if not header:
        return ""
    return str(header).strip().lower().replace("_", " ")


def import_excel(filepath, sheet_name=None):
    """Import billing records from an Excel file.

    Args:
        filepath: Path to .xlsx file
        sheet_name: Sheet to read (default: first sheet or 'Current')

    Returns:
        dict with keys: imported, skipped, errors, total_rows
    """
    result = {"imported": 0, "skipped": 0, "errors": [], "total_rows": 0, "filename": os.path.basename(filepath)}

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        result["errors"].append(f"Cannot open workbook: {e}")
        return result

    # Select sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "Current" in wb.sheetnames:
        ws = wb["Current"]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result["errors"].append("Empty sheet")
        return result

    # Map headers (hardcoded + learned)
    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    col_map, unmapped = enhance_column_map(raw_headers, COLUMN_MAP, source_format="EXCEL")

    if not col_map:
        result["errors"].append(f"No recognized columns. Headers found: {raw_headers[:10]}")
        return result

    if unmapped:
        result["unmapped_columns"] = [u["header"] for u in unmapped]

    # Build dedup set from existing records
    existing = build_dedup_set()

    data_rows = rows[1:]
    result["total_rows"] = len(data_rows)
    batch = []
    batch_size = 500

    for row_idx, row in enumerate(data_rows, start=2):
        try:
            record_data = {}
            for col_idx, field_name in col_map.items():
                if col_idx < len(row):
                    record_data[field_name] = row[col_idx]

            # Required fields
            patient_name = str(record_data.get("patient_name", "")).strip()
            if not patient_name:
                continue

            service_date = parse_date(record_data.get("service_date"))
            if not service_date:
                result["skipped"] += 1
                continue

            scan_type = str(record_data.get("scan_type", "")).strip() or "UNKNOWN"
            modality = enhanced_normalize_modality(record_data.get("modality"))
            referring_doctor = str(record_data.get("referring_doctor", "")).strip() or "UNKNOWN"

            # Dedup check
            if is_duplicate(patient_name, service_date, scan_type, modality, existing):
                result["skipped"] += 1
                continue

            description = str(record_data.get("description", "")).strip() if record_data.get("description") else None

            primary = parse_float(record_data.get("primary_payment"))
            secondary = parse_float(record_data.get("secondary_payment"))
            extra = parse_float(record_data.get("extra_charges"))
            total = parse_float(record_data.get("total_payment"))
            total = compute_total_payment(primary, secondary, total, extra)

            rec = BillingRecord(
                patient_name=patient_name,
                referring_doctor=referring_doctor,
                scan_type=scan_type,
                gado_used=parse_bool(record_data.get("gado_used")),
                insurance_carrier=enhanced_normalize_carrier(record_data.get("insurance_carrier")),
                modality=modality,
                service_date=service_date,
                primary_payment=primary,
                secondary_payment=secondary,
                total_payment=total,
                extra_charges=extra,
                reading_physician=str(record_data.get("reading_physician", "")).strip() or None,
                patient_id=int(parse_float(record_data.get("patient_id"))) if record_data.get("patient_id") else None,
                description=description,
                is_psma=detect_psma(description, scan_type),
                import_source="EXCEL_IMPORT",
            )
            batch.append(rec)

            if len(batch) >= batch_size:
                db.session.bulk_save_objects(batch)
                db.session.commit()
                result["imported"] += len(batch)
                batch = []

        except Exception as e:
            result["errors"].append(f"Row {row_idx}: {e}")

    # Final batch
    if batch:
        db.session.bulk_save_objects(batch)
        db.session.commit()
        result["imported"] += len(batch)

    wb.close()
    return result
